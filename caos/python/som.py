from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "=SUM(5, 4)"
print(ws["A1"].value)
print(type(ws))